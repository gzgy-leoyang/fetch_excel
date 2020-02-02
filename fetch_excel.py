import openpyxl
from openpyxl import Workbook

import os
import sys
import getopt

# 逐一打印单元值 
# for row in sheet.rows:
#     for cell in row:
#         print( cell.value)
# 等于如下,
# 直接构造 list
# for row in rows:
#     line = [ col.value for col in row ]

#######################3
# @berif 提取列数据
# file_name               文件
# page_index           表索引
# line_index      行列索引
# line_mode，行列模式
def fetch_line( file_name ,page_index ,line_index,line_mode):
    global start_line_index
    max_limit = 0       # 标记行/列 模式下的最大范围
    
    wb = openpyxl.load_workbook(cur_path+'/'+file_name)
    sheet_list = wb.get_sheet_names()

    if len( sheet_list ) > (page_index-1) :
        sheet = wb.get_sheet_by_name( sheet_list[page_index-1] )
    else :
        print ( ' 异常：指定表索引超出范围' )
        return None

    # 行列的最大访问范围
    if line_mode == 'row':
        max_limit = sheet.max_row
    elif line_mode == 'col':
        max_limit = sheet.max_column
    else:
        line_mode == 'row'
        max_limit = sheet.max_row

    if max_limit  >= line_index :
        if line_mode == 'row':
            rows = sheet.rows
            cnt = 0 
            for row in rows:
                cnt = cnt + 1
                if cnt == line_index:
                    return [ col.value for col in row ]            
        elif line_mode == 'col':
            columns = sheet.columns
            cnt = 0 
            for column in columns:
                cnt = cnt + 1
                if cnt == line_index:
                    return [ row.value for row in column ]
    else :
        print('异常：指定行索引超出范围')
        return None

##########################
# @berif 将输入行列的内容追加到输出文件
# sheet : 指定表
# line_context，添加内容
# line_mode,添加模式，行/列
def append_line(sheet,line_context,line_mode):
    if line_context != None:
        if line_mode == 'row':
            # 按行模式添加
            row = sheet.max_row + 1
            for i in  range(  len( line_context )  ):
                sheet.cell( row ,i+1).value = line_context[i]
        elif line_mode == 'col':
            # 按列模式添加
            column = sheet.max_column + 1
            for i in  range(  len( line_context )  ):
                sheet.cell( i+1,column).value = line_context[i]
        else:
            print (' 写入模式无效 ')

############################
# @berif 打印 help 信息
def usage( ):
    print(' Usage: python3 fetch_excel [-s start-index] [-e end-index] [-m row/col] [-o *.xlsx] [-t title-index] [-p page-index]')
    print(' options:')    
    print(' -s  提取行或列的起始位置，e.g. 1,2,3,..., def: 1')
    print(' -e  提取行或列的结束位置，e.g. 1,2,3,..., def: 1')
    print(' -o  指定汇总输出文件，e.g. my-out.xlsx ，def: out.xlsx')
    print(' -p  指定输入文件中的表序号，文件中第一张表对应为1，e.g. 1,2,3,..., def: 1')
    print(' -t  指定提取的标题行/列的位置,标题仅提取一次，e.g. 1,2,3,..., def: 1')
    print(' -m  指定提取模式，分为 row 和 col 模式, def : row')
    print(' Fetch Excel v1.0.0  2020/2/2 ( leoyang20102013@163.com )')

#################
## <<  程序入口 >> ##
opts,args = getopt.getopt( sys.argv[1:] ,'s:e:t:p:o:m:h')
start_line_index = 1
end_line_index = 1
fetch_sheet_index = 1
title_line_index = 1
fetch_line_mode = 'row'
out_file_name = 'out.xlsx'

for op,value in opts:
    if op == '-s':
        if value.isdigit():
            start_line_index = int(value)
        else :
            print('起始行号：无效字符')
            usage()
            sys.exit()
    elif op == '-e':
        if value.isdigit():
            end_line_index = int(value)
        else :
            print('结束行号：无效字符')
            usage()
            sys.exit()
    elif op == '-t':
        if value.isdigit():
            title_line_index = int(value)
        else :
            print('标题行：无效字符')
            usage()
            sys.exit()
    elif op == '-p':
        if value.isdigit():
            fetch_sheet_index = int(value)
        else :
            print('表索引：无效字符')
            usage()
            sys.exit()
    elif op == '-o':
        if value.find('.xlsx') > 0 :
            out_file_name = value
        else :
            print ( '输出文件：无效格式 ')
            usage()
            sys.exit()    
    elif op == '-m':
        if value == 'row' or value == 'col' :
            fetch_line_mode = value
        else:
            print ( ' 行列模式：无效参数 ')
            usage()
            sys.exit()
    elif op == '-h':
        usage()
        sys.exit()

print( '表: %d  标题: %d' % (fetch_sheet_index,title_line_index,) )
print( '模式：%s    起始行/列: %d  结束行/列: %d' % (fetch_line_mode,start_line_index,end_line_index) )
print( '输出: %s' % (out_file_name) )

if input('回车键开始，任意键退出...') !=  "":
    sys.exit()

## 查询是否有文件，如果有该文件，执行删除，再重新建同名文件  
cur_path  = sys.path[0]
file_name  = cur_path+'/'+out_file_name
if os.access( file_name , os.F_OK ):
    print ( 'File exist, remove it' )
    os.remove(file_name)

## 新建文件
w_wb = Workbook()
w_sheet_list = w_wb.get_sheet_names()
w_wb_sheet = w_wb.get_sheet_by_name(w_sheet_list[0])
w_wb_sheet = w_wb.active
w_wb.save(file_name)
print ( 'Create file %s , %s' % (file_name,w_sheet_list[0]) )

# 遍历全部文件
file_list = os.listdir(cur_path)
for file in file_list :
    # 遍历文件的范围：排除输出文件，仅检查xlsx文件，且不是 .~*.xlsx 临时文件
    if (( file != out_file_name )  and  ( file[-4:] == 'xlsx' ) and ( not file[:2] == '.~' )):
        print("Fetching....",file)
         # 从 file 中提取指定内容
         # 提取标题栏，，加入输出文件（仅执行一次）
        if title_line_index > 0 :
            line_context = fetch_line( file , fetch_sheet_index ,title_line_index ,fetch_line_mode)
            append_line( w_wb_sheet ,line_context ,fetch_line_mode )
            title_line_index = 0
        # 提取指定行，加入输出文件
        if end_line_index > start_line_index :
            # 多行提取
            for i in range( 0,end_line_index - start_line_index + 1):
                line_context = fetch_line( file , fetch_sheet_index , (start_line_index + int(i)) ,fetch_line_mode)
                append_line( w_wb_sheet , line_context  ,fetch_line_mode )
        else :
            # 单行提取,没有设置 target_row1，则默认为1，小于等于 start_line_index
            line_context = fetch_line( file , fetch_sheet_index , start_line_index ,fetch_line_mode)
            append_line( w_wb_sheet , line_context ,fetch_line_mode )
w_wb.save( file_name )